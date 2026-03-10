"""
Analyze K-choose-N evaluation results: precision, recall, exact match, and breakdowns.
"""
import json
from pathlib import Path
from collections import defaultdict

import yaml


def load_config() -> dict:
    config_path = Path(__file__).parent.parent.parent / "config.yaml"
    with open(config_path, "r") as f:
        return yaml.safe_load(f)


def _avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _compute_metrics(results: list[dict]) -> dict:
    if not results:
        return {"precision": 0.0, "recall": 0.0, "exact_match": 0.0, "count": 0}
    return {
        "precision": _avg([r["precision"] for r in results]),
        "recall": _avg([r["recall"] for r in results]),
        "exact_match": sum(1 for r in results if r["exact_match"]) / len(results),
        "count": len(results),
    }


def _bucket_by_comment_count(eval_users: list[dict], results: list[dict]) -> dict:
    """Bucket users by activity level."""
    user_comment_count = {u["user_name"]: u["num_remaining_comments"] for u in eval_users}
    buckets = {
        "5-10": (5, 10),
        "11-20": (11, 20),
        "21-50": (21, 50),
        "51+": (51, float("inf")),
    }
    bucket_results = {name: [] for name in buckets}
    for r in results:
        count = user_comment_count.get(r["user_name"], 0)
        for name, (lo, hi) in buckets.items():
            if lo <= count <= hi:
                bucket_results[name].append(r)
                break
    return bucket_results


def _print_metrics(label: str, metrics: dict, k: int, n: int):
    print(f"  {label}")
    print(f"    Queries:      {metrics['count']}")
    print(f"    Precision:    {metrics['precision']:.4f}")
    print(f"    Recall:       {metrics['recall']:.4f}")
    print(f"    Exact Match:  {metrics['exact_match']:.4f}")


def main():
    config = load_config()
    base_dir = Path(__file__).parent.parent.parent / "data"
    eval_dir = base_dir / "evaluation"

    with open(eval_dir / "eval_results.json", "r") as f:
        results = json.load(f)

    with open(eval_dir / "eval_users.json", "r") as f:
        eval_users = json.load(f)

    with open(base_dir / "processed" / "all_posts_meta.json", "r") as f:
        all_posts_meta = json.load(f)

    posts_by_conv = {p["conversation_id"]: p for p in all_posts_meta}
    k = results[0]["k"] if results else 10
    n = results[0]["n"] if results else 3

    print("=" * 60)
    print("EVALUATION RESULTS ANALYSIS")
    print("=" * 60)
    print(f"Total queries: {len(results)}")
    print(f"K = {k}, N = {n}, random baseline precision = {n/k:.3f}")
    print()

    # --- Overall ---
    overall = _compute_metrics(results)
    _print_metrics("Overall", overall, k, n)
    print()

    # --- Breakdown by comment count buckets ---
    print("-" * 40)
    print("Metrics by activity level (comment count):")
    bucket_results = _bucket_by_comment_count(eval_users, results)
    for bucket_name, b_results in bucket_results.items():
        if b_results:
            metrics = _compute_metrics(b_results)
            _print_metrics(bucket_name, metrics, k, n)
        else:
            print(f"  {bucket_name:>8s}: (no data)")

    # --- Breakdown by subreddit (using first positive post) ---
    print()
    print("-" * 40)
    print("Metrics by subreddit (first positive post's subreddit):")
    sub_results = defaultdict(list)
    for r in results:
        pos_ids = r.get("positive_conversation_ids", [])
        if pos_ids:
            post = posts_by_conv.get(pos_ids[0])
            sub = post["subreddit"] if post else "unknown"
        else:
            sub = "unknown"
        sub_results[sub].append(r)
    for sub in sorted(sub_results.keys()):
        sr = sub_results[sub]
        metrics = _compute_metrics(sr)
        _print_metrics(sub, metrics, k, n)

    # --- Export to Excel ---
    try:
        import openpyxl
        _export_excel(results, eval_users, posts_by_conv, eval_dir, k, n)
    except ImportError:
        print("\nopenpyxl not installed — skipping Excel export.")


def _export_excel(results, eval_users, posts_by_conv, eval_dir, k, n):
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    overall = _compute_metrics(results)
    ws.append(["K", "N", "Total Queries", "Precision", "Recall",
               "Exact Match", "Random Baseline Precision"])
    ws.append([
        k, n, overall["count"],
        overall["precision"], overall["recall"], overall["exact_match"],
        n / k,
    ])

    # Sheet 2: Per-query results
    ws2 = wb.create_sheet("Per-Query Results")
    headers = ["user_name", "positive_conversation_ids", "k", "n",
               "gold_indices", "chosen_indices", "precision", "recall",
               "exact_match", "reason"]
    ws2.append(headers)
    for r in results:
        row = []
        for h in headers:
            val = r.get(h, "")
            if isinstance(val, (list, set)):
                val = json.dumps(val)
            row.append(val)
        ws2.append(row)

    # Sheet 3: By activity bucket
    ws3 = wb.create_sheet("By Activity")
    ws3.append(["Bucket", "N_queries", "Precision", "Recall", "Exact Match"])
    bucket_results = _bucket_by_comment_count(eval_users, results)
    for bucket_name, br in bucket_results.items():
        if br:
            m = _compute_metrics(br)
            ws3.append([bucket_name, m["count"], m["precision"],
                        m["recall"], m["exact_match"]])

    # Sheet 4: By subreddit
    ws4 = wb.create_sheet("By Subreddit")
    ws4.append(["Subreddit", "N_queries", "Precision", "Recall", "Exact Match"])
    sub_results = defaultdict(list)
    for r in results:
        pos_ids = r.get("positive_conversation_ids", [])
        if pos_ids:
            post = posts_by_conv.get(pos_ids[0])
            sub = post["subreddit"] if post else "unknown"
        else:
            sub = "unknown"
        sub_results[sub].append(r)
    for sub in sorted(sub_results.keys()):
        br = sub_results[sub]
        m = _compute_metrics(br)
        ws4.append([sub, m["count"], m["precision"], m["recall"], m["exact_match"]])

    output_path = eval_dir / "eval_analysis.xlsx"
    wb.save(output_path)
    print(f"\nExported analysis to {output_path}")


if __name__ == "__main__":
    main()
